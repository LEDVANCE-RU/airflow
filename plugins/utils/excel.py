from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Alignment, Border, Font


class ExcelPrettifier:
    def __init__(self, filepath: str, autosave: bool = False):
        self.filepath = filepath
        self.autosave = autosave
        self.wb = None
        self.ws = None

    def __enter__(self):
        self.wb = load_workbook(self.filepath)
        self.ws = self.wb.active
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        if self.autosave:
            self.wb.save(self.filepath)
        self.wb.close()

    def merge_cells_by_unique_row_values(self, header_index: int, column_indices: list[int]):
        start_row = header_index + 1
        current_row_values = None
        merge_start = start_row

        for row in range(start_row, self.ws.max_row + 2):
            row_values = tuple(self.ws.cell(row=row, column=col_idx).value for col_idx in column_indices)
            if current_row_values and current_row_values != row_values:
                if merge_start < row - 1:
                    for col_idx in column_indices:
                        self.ws.merge_cells(
                            start_row=merge_start, start_column=col_idx,
                            end_row=row - 1, end_column=col_idx
                        )
                merge_start = row
            current_row_values = row_values


    def apply_style(self,
                    ranges: list[(int, int, int, int)] = None,
                    border_style: Border = None,
                    alignment: Alignment = None,
                    font: Font = None):
        max_col = self.ws.max_column
        max_row = self.ws.max_row
        if max_col == 1 and max_row == 1 and self.ws.cell(1, 1).value is None:
            return
        if not ranges:
            ranges = [(1, 1, max_col, max_row)]
        for range_ in ranges:
            if len(range_) != 4:
                raise ValueError(f"Incorrect range '{range_}', expected tuple of 4 values.")
            start_col, start_row = tuple(max(r, 1) if r else 1 for r in range_[0:2])
            end_col = min(range_[2] or max_col, max_col)
            end_row = min(range_[3] or max_row, max_row)
            for row in self.ws.iter_rows(min_col=start_col, min_row=start_row, max_col=end_col, max_row=end_row):
                for cell in row:
                    if border_style is not None:
                        cell.border = border_style
                    if alignment is not None:
                        cell.alignment = alignment
                    if font is not None:
                        cell.font = font


    def auto_fit_columns_advanced(self, min_width=2, max_width=50, padding=2):
        column_widths = {}

        for row in self.ws.iter_rows():
            for cell in row:
                if cell.value:
                    content = str(cell.value)

                    if '\n' in content:
                        lines = content.split('\n')
                        max_line_length = max(len(line) for line in lines)
                        content_length = max_line_length
                    else:
                        content_length = len(content)

                    if cell.font:
                        content_length = content_length * cell.font.sz / 10
                        if cell.font.bold:
                            content_length *= 1.1

                    current_max = column_widths.get(cell.column_letter, 0)
                    column_widths[cell.column_letter] = max(current_max, content_length)

        for column_letter, content_length in column_widths.items():
            width = min(max(content_length + padding, min_width), max_width)
            self.ws.column_dimensions[column_letter].width = width
